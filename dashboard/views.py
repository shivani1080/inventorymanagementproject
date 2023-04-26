from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import Product,Issued_Items
from django.contrib.auth.decorators import login_required
from .forms import ProductForm,orderform,sendemailform
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings 

from reportlab.pdfgen import canvas


from django.http import HttpResponse
from openpyxl import Workbook
from .models import Product, Issued_Items
# Create your views here.

@login_required
def index(request):
    issueditems=Issued_Items.objects.all()
    product=Product.objects.all()
    items_count = issueditems.count()
    product_count = product.count()
    workers_count =User.objects.all().count()
    
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    if request.method=='POST':
         form=orderform(request.POST)
         if form.is_valid():
            instance=form.save(commit=False)
            instance.staff=request.user
            instance.save()
            return redirect('dashboardindex')
    else:
        form=orderform()
    context={
        'issueditems':issueditems,
        'form':form,
        'product':product,
        'product_count': product_count,
        'workers_count': workers_count,
        'items_count': items_count,
        'emailform':emailform,

    }
    return render(request,'dashboard/index.html',context)
@login_required
def staff(request):
    workers=User.objects.all()
    workers_count = workers.count()
    items_count =Issued_Items.objects.all().count()
    product_count =Product.objects.all().count()
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    context= {
        'workers':workers,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
        'emailform':emailform,
    }
    return render(request,'dashboard/staff.html',context)
@login_required
def staff_detail(request,pk):
    worker=User.objects.all()
    workers_count = worker.count()
    items_count =Issued_Items.objects.all().count()
    product_count =Product.objects.all().count()
    if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
    else:
        emailform=sendemailform()
    workers=User.objects.get(id=pk)
    context={
        'workers':workers,
        'emailform':emailform,
        'worker':worker,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
    }

    return render(request,'dashboard/staff_detail.html',context)
@login_required
def product(request):
    items=Product.objects.all()
    product_count = items.count()
    #items=Product.objects.raw('SELECT * FROM dashboard_product')
    workers_count =User.objects.all().count()
    items_count =Issued_Items.objects.all().count()

    if request.method=='POST':
        form=ProductForm(request.POST)
        if form.is_valid():
            form.save()
            product_name = form.cleaned_data.get('name')
            messages.success(request, f'{product_name} has been added')
            return redirect('dashboardproduct')
    else:
        form=ProductForm()
    context={
        'items':items,
        'form':form,
        'workers_count': workers_count,
        'items_count': items_count,
        'product_count': product_count,
        
    }
    return render(request,'dashboard/product.html',context)   
@login_required
def product_delete(request,pk):
    item=Product.objects.get(id=pk)
    if request.method=='POST':
        item.delete()
        return redirect('dashboardproduct')
    return render(request,'dashboard/product_delete.html') 
@login_required
def product_update(request,pk):
    item=Product.objects.get(id=pk)
    if request.method=='POST':
        form=ProductForm(request.POST,instance=item)
        if form.is_valid():
            form.save()
            return redirect('dashboardproduct')
    else:
        form=ProductForm(instance=item)
    context={
        'form':form

    }
    return render(request,'dashboard/product_update.html',context)

@login_required
def issued_items(request):
 items=Issued_Items.objects.all()
 items_count = items.count()
 workers_count =User.objects.all().count()
 product_count =Product.objects.all().count()
 if request.method=='POST':
        emailform=sendemailform(request.POST)
        if emailform.is_valid():
            name=emailform.cleaned_data['title']
            toemail=emailform.cleaned_data['empemail']
            message=emailform.cleaned_data['message']
            send_mail(name,message,'settings.EMAIL_HOST_USER',[toemail],fail_silently=False)
            return redirect('dashboardindex')
 else:
    emailform=sendemailform()
 context={
     'items':items,
     'workers_count': workers_count,
     'items_count': items_count,
     'product_count': product_count,
     'emailform':emailform,
 }

 return render(request,'dashboard/issueditems.html',context)

def generate_pdf(request):
    # Get all products and issued items
    products = Product.objects.all()
    issued_items = Issued_Items.objects.all()

    # Create a response object with PDF content type
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="products.pdf"'

    # Create a canvas object and draw the products and issued items
    p = canvas.Canvas(response)
    p.drawString(100, 750, "List of Products")
    p.drawString(100, 700, "------------------------")
    y = 650
    for product in products:
        p.drawString(100, y, f"{product.name} - {product.quantity}")
        y -= 20
    p.showPage()

    p.drawString(100, 750, "List of Issued Items")
    p.drawString(100, 700, "------------------------")
    y = 650
    for issued_item in issued_items:
        p.drawString(100, y, f"{issued_item.product} issued to {issued_item.staff}")
        y -= 20
    p.showPage()

    p.save()
    return response


def generate_excel_file(request):
    # Create a new workbook and add a worksheet
    wb = Workbook()
    ws_products = wb.active
    ws_products.title = "Products"
    ws_issued_items = wb.create_sheet("Issued Items")

    # Write headers to worksheet for Products
    ws_products['A1'] = 'Asset'
    ws_products['B1'] = 'SNO'
    ws_products['C1'] = 'Name'
    ws_products['D1'] = 'Category'
    ws_products['E1'] = 'Quantity'
    ws_products['F1'] = 'Model'
    ws_products['G1'] = 'Price'

    # Write headers to worksheet for Issued_Items
    ws_issued_items['A1'] = 'Product'
    ws_issued_items['B1'] = 'Issued to'
    ws_issued_items['C1'] = 'Quantity'
    ws_issued_items['D1'] = 'Location'

    # Write data to Products worksheet
    products = Product.objects.all()
    for row, product in enumerate(products, start=2):
        
        ws_products.cell(row=row, column=1, value=product.asset)
        ws_products.cell(row=row, column=2, value=product.sno)
        ws_products.cell(row=row, column=3, value=product.name)
        ws_products.cell(row=row, column=4, value=product.category)
        ws_products.cell(row=row, column=5, value=product.quantity)
        ws_products.cell(row=row, column=6, value=product.model)
        ws_products.cell(row=row, column=7, value=product.price)

    # Write data to Issued_Items worksheet
    issued_items = Issued_Items.objects.all()
    for row, issued_item in enumerate(issued_items, start=2):
        
        ws_issued_items.cell(row=row, column=1, value=issued_item.product.name)
        ws_issued_items.cell(row=row, column=2, value=issued_item.staff.username)
        ws_issued_items.cell(row=row, column=3, value=issued_item.issueditem_quantity)
        ws_issued_items.cell(row=row, column=4, value=issued_item.location)

    # Set the response headers
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'

    # Save the workbook to the response
    wb.save(response)
    
    

    return response
